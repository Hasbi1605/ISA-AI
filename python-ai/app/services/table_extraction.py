from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

import pdfplumber
from docx import Document as DocxDocument
from openpyxl import load_workbook


def _clean_cell_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _compact_rows(rows: list[list[Any]]) -> list[list[str]]:
    cleaned_rows: list[list[str]] = []

    for row in rows:
        cleaned_row = [_clean_cell_text(cell) for cell in row]
        if any(cell for cell in cleaned_row):
            cleaned_rows.append(cleaned_row)

    return cleaned_rows


def _table_payload(rows: list[list[str]], source: str, page: int | None = None) -> dict[str, Any]:
    if not rows:
        return {
            "source": source,
            "page": page,
            "header": [],
            "rows": [],
        }

    header = rows[0] if any(rows[0]) else []
    body_rows = rows[1:] if header else rows

    return {
        "source": source,
        "page": page,
        "header": header,
        "rows": body_rows,
    }


def extract_tables_from_pdf(file_path: str) -> list[dict[str, Any]]:
    tables: list[dict[str, Any]] = []

    with pdfplumber.open(file_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            for raw_table in page.extract_tables() or []:
                rows = _compact_rows(raw_table)
                if not rows:
                    continue

                tables.append(_table_payload(rows, source="pdf", page=page_number))

    return tables


def extract_tables_from_docx(file_path: str) -> list[dict[str, Any]]:
    document = DocxDocument(file_path)
    tables: list[dict[str, Any]] = []

    for index, table in enumerate(document.tables, start=1):
        rows = []
        for row in table.rows:
            rows.append([_clean_cell_text(cell.text) for cell in row.cells])

        rows = _compact_rows(rows)
        if not rows:
            continue

        payload = _table_payload(rows, source="docx", page=None)
        payload["table_index"] = index
        tables.append(payload)

    return tables


def extract_tables_from_xlsx(file_path: str) -> list[dict[str, Any]]:
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    tables: list[dict[str, Any]] = []

    try:
        for sheet in workbook.worksheets:
            rows = _compact_rows([
                [cell for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ])

            if not rows:
                continue

            payload = _table_payload(rows, source="xlsx", page=None)
            payload["sheet"] = sheet.title
            tables.append(payload)
    finally:
        workbook.close()

    return tables


def extract_tables_from_csv(file_path: str) -> list[dict[str, Any]]:
    rows: list[list[str]] = []

    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(file_path, newline="", encoding=encoding) as handle:
                sample = handle.read(4096)
                handle.seek(0)

                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel

                rows = _compact_rows([row for row in csv.reader(handle, dialect)])
            break
        except UnicodeDecodeError:
            continue

    return [_table_payload(rows, source="csv", page=None)] if rows else []


def extract_tables_from_file(file_path: str) -> list[dict[str, Any]]:
    suffix = Path(file_path).suffix.lower()

    if suffix == ".pdf":
        return extract_tables_from_pdf(file_path)

    if suffix == ".docx":
        return extract_tables_from_docx(file_path)

    if suffix == ".xlsx":
        return extract_tables_from_xlsx(file_path)

    if suffix == ".csv":
        return extract_tables_from_csv(file_path)

    raise ValueError(f"Unsupported file type for table extraction: {suffix or 'unknown'}")
