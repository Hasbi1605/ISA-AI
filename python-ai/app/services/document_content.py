from __future__ import annotations

import csv
import re
from html import escape
from pathlib import Path
from typing import Any

import pdfplumber
from docx import Document as DocxDocument
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table as DocxTable
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook


def _clean_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _paragraph_html(text: str) -> str:
    cleaned = _clean_text(text)

    if cleaned == "":
        return ""

    return f"<p>{escape(cleaned)}</p>"


def _table_html(rows: list[list[Any]]) -> str:
    cleaned_rows = [
        [_clean_text(cell) for cell in row]
        for row in rows
    ]
    cleaned_rows = [row for row in cleaned_rows if any(row)]

    if not cleaned_rows:
        return ""

    header = cleaned_rows[0]
    body_rows = cleaned_rows[1:] if any(header) else cleaned_rows
    header_html = ""

    if any(header):
        header_html = "<thead><tr>{}</tr></thead>".format(
            "".join(f"<th>{escape(cell)}</th>" for cell in header)
        )

    body_html = "".join(
        "<tr>{}</tr>".format("".join(f"<td>{escape(cell)}</td>" for cell in row))
        for row in body_rows
    )

    return f"<table>{header_html}<tbody>{body_html}</tbody></table>"


def _wrap_document(title: str, sections: list[str]) -> str:
    content = "\n".join(section for section in sections if section.strip())

    if not content:
        content = "<p>Tidak ada teks yang dapat diekstrak dari dokumen ini.</p>"

    return f"<article><h1>{escape(title)}</h1>{content}</article>"


def _extract_pdf_content_html(file_path: str, title: str) -> str:
    sections: list[str] = []

    with pdfplumber.open(file_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            page_sections: list[str] = []
            text = page.extract_text() or ""

            for block in re.split(r"\n{2,}", text):
                lines = [_clean_text(line) for line in block.splitlines()]
                paragraph = " ".join(line for line in lines if line)
                paragraph_html = _paragraph_html(paragraph)
                if paragraph_html:
                    page_sections.append(paragraph_html)

            for raw_table in page.extract_tables() or []:
                table_html = _table_html(raw_table)
                if table_html:
                    page_sections.append(table_html)

            if page_sections:
                sections.append(
                    f"<section><h2>Halaman {page_number}</h2>{''.join(page_sections)}</section>"
                )

    return _wrap_document(title, sections)


def _extract_docx_content_html(file_path: str, title: str) -> str:
    document = DocxDocument(file_path)
    sections: list[str] = []

    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            paragraph = Paragraph(child, document)
            paragraph_html = _paragraph_html(paragraph.text)
            if paragraph_html:
                sections.append(paragraph_html)
        elif isinstance(child, CT_Tbl):
            table = DocxTable(child, document)
            rows = [[cell.text for cell in row.cells] for row in table.rows]
            table_html = _table_html(rows)
            if table_html:
                sections.append(table_html)

    return _wrap_document(title, sections)


def _extract_xlsx_content_html(file_path: str, title: str) -> str:
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sections: list[str] = []

    for sheet in workbook.worksheets:
        rows = [
            [cell for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        table_html = _table_html(rows)

        if table_html:
            sections.append(f"<section><h2>{escape(sheet.title)}</h2>{table_html}</section>")

    workbook.close()

    return _wrap_document(title, sections)


def _extract_csv_content_html(file_path: str, title: str) -> str:
    rows: list[list[str]] = []

    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(file_path, newline="", encoding=encoding) as handle:
                sample = handle.read(4096)
                handle.seek(0)

                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel

                rows = [row for row in csv.reader(handle, dialect)]
            break
        except UnicodeDecodeError:
            continue

    table_html = _table_html(rows)
    sections = [f"<section><h2>{escape(title)}</h2>{table_html}</section>"] if table_html else []

    return _wrap_document(title, sections)


def extract_document_content_html(file_path: str, filename: str | None = None) -> str:
    suffix = Path(file_path).suffix.lower()
    title = filename or Path(file_path).name

    if suffix == ".pdf":
        return _extract_pdf_content_html(file_path, title)

    if suffix == ".docx":
        return _extract_docx_content_html(file_path, title)

    if suffix == ".xlsx":
        return _extract_xlsx_content_html(file_path, title)

    if suffix == ".csv":
        return _extract_csv_content_html(file_path, title)

    raise ValueError(f"Unsupported file type for content extraction: {suffix or 'unknown'}")
