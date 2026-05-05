import os
import sys
from io import BytesIO

from docx import Document as DocxDocument
from openpyxl import Workbook
from weasyprint import HTML

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.services.table_extraction import extract_tables_from_file


def _write_pdf_fixture(path):
    html = """
    <html>
      <head>
        <style>
          table { border-collapse: collapse; width: 100%; }
          th, td { border: 1px solid #111111; padding: 8px; }
        </style>
      </head>
      <body>
        <table>
          <tr><th>Nama</th><th>Nilai</th></tr>
          <tr><td>A</td><td>10</td></tr>
          <tr><td>B</td><td>20</td></tr>
        </table>
      </body>
    </html>
    """
    HTML(string=html).write_pdf(target=str(path))


def _write_docx_fixture(path):
    document = DocxDocument()
    table = document.add_table(rows=3, cols=2)
    table.cell(0, 0).text = "Nama"
    table.cell(0, 1).text = "Nilai"
    table.cell(1, 0).text = "A"
    table.cell(1, 1).text = "10"
    table.cell(2, 0).text = "B"
    table.cell(2, 1).text = "20"
    document.save(path)


def _write_xlsx_fixture(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anggaran"
    sheet.append(["Nama", "Nilai"])
    sheet.append(["A", 10])
    sheet.append(["B", 20])

    second_sheet = workbook.create_sheet("Ringkasan")
    second_sheet.append(["Kategori", "Total"])
    second_sheet.append(["Operasional", 30])

    workbook.save(path)


def test_extract_tables_from_pdf(tmp_path):
    pdf_path = tmp_path / "sample.pdf"
    _write_pdf_fixture(pdf_path)

    tables = extract_tables_from_file(str(pdf_path))

    assert len(tables) >= 1
    assert tables[0]["header"] == ["Nama", "Nilai"]
    assert tables[0]["rows"][0] == ["A", "10"]


def test_extract_tables_from_docx(tmp_path):
    docx_path = tmp_path / "sample.docx"
    _write_docx_fixture(docx_path)

    tables = extract_tables_from_file(str(docx_path))

    assert len(tables) == 1
    assert tables[0]["header"] == ["Nama", "Nilai"]
    assert tables[0]["rows"][1] == ["B", "20"]


def test_extract_tables_from_xlsx(tmp_path):
    xlsx_path = tmp_path / "sample.xlsx"
    _write_xlsx_fixture(xlsx_path)

    tables = extract_tables_from_file(str(xlsx_path))

    assert len(tables) == 2
    assert tables[0]["sheet"] == "Anggaran"
    assert tables[0]["header"] == ["Nama", "Nilai"]
    assert tables[0]["rows"][0] == ["A", "10"]
    assert tables[1]["sheet"] == "Ringkasan"
    assert tables[1]["rows"][0] == ["Operasional", "30"]


def test_extract_tables_from_csv(tmp_path):
    csv_path = tmp_path / "sample.csv"
    csv_path.write_text("Nama,Nilai\nA,10\nB,20\n", encoding="utf-8")

    tables = extract_tables_from_file(str(csv_path))

    assert len(tables) == 1
    assert tables[0]["header"] == ["Nama", "Nilai"]
    assert tables[0]["rows"][1] == ["B", "20"]
